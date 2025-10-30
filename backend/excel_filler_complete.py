import json
import os
from openpyxl import load_workbook

def fill_itr_excel(form16_json_path, aadhar_json_path, passbook_json_path, excel_file_path, output_file_path):
    # Load JSON data
    with open(form16_json_path, 'r') as f:
        form16_data = json.load(f)
    
    with open(aadhar_json_path, 'r') as f:
        aadhar_data = json.load(f)
    
    with open(passbook_json_path, 'r') as f:
        passbook_data = json.load(f)
    
    # Load Excel workbook
    wb = load_workbook(excel_file_path)
    ws = wb.active
    
    # Form 16 cell mapping
    form16_mapping = {
        "pan": "AN7",
        "employee_address": "O11",
        "gross_salary": "AO35",
        "salary_section_17_1": "AO36",
        "prerequisites_section_17_2": "AO37",
        "profits_section_17_3": "AO38",
        "total_exemption_section_10": "AO45",
        "standard_deduction_16_ia": "AO54",
        "entertainment_allowance_16_ii": "AO55",
        "tax_on_employment_16_iii": "AO56",
        "income_chargeable_salaries": "AO57",
        "gross_total_income": "AO93",
        "tax_on_total_income": "AO140",
        "rebate_87A": "AO141",
        "health_education_cess": "AO144",
        "relief_section_89": "AO146"
    }
    
    # Aadhar cell mapping
    aadhar_mapping = {
        "aadhar_number": "AN8",
        "dob": "AN11",
        "pin_code": "AA15",
        "flat_door_block_no": "E11",
        "building_premises_village": "O11",
        "road_street_post_office": "E13",
        "area_locality": "W13",
        "town_city_district": "AN13",
        "state": "E15"
    }
    
    # Parse name from passbook
    def parse_name(full_name):
        if not full_name:
            return "", "", ""
        
        # Remove common prefixes and clean
        name_parts = full_name.replace("MR ", "").replace("MRS ", "").replace("MS ", "").strip().split()
        
        if len(name_parts) == 1:
            return name_parts[0], "", ""
        elif len(name_parts) == 2:
            return name_parts[0], "", name_parts[1]
        else:
            return name_parts[0], " ".join(name_parts[1:-1]), name_parts[-1]
    
    # Passbook cell mapping
    passbook_mapping = {
    }
    
    # Deductions that need dual cells
    dual_cell_mapping = {
        "deduction_80C": ["AB96", "AN96"],
        "deduction_80CCC": ["AB97", "AN97"],
        "deduction_80CCD1": ["AB98", "AN98"],
        "deduction_80CCD1B": ["AB99", "AN99"],
        "deduction_80CCD2": ["AB101", "AN101"],
        "deduction_80D": ["AB102", "AN102"],
        "deduction_80E": ["AB111", "AN111"],
        "deduction_80G": ["AB115", "AN115"],
        "deduction_80TTA": ["AB122", "AN122"]
    }
    
    # Fill Form 16 data
    for json_key, cell_address in form16_mapping.items():
        if json_key in form16_data:
            ws[cell_address] = form16_data[json_key]
    
    # Fill Aadhar data
    for json_key, cell_address in aadhar_mapping.items():
        if json_key in aadhar_data:
            ws[cell_address] = aadhar_data[json_key]
    
    # Fill passbook data
    for json_key, cell_address in passbook_mapping.items():
        if json_key in passbook_data:
            ws[cell_address] = passbook_data[json_key]
    
    # Fill parsed name from passbook
    if "name" in passbook_data:
        first_name, middle_name, last_name = parse_name(passbook_data["name"])
        ws["E7"] = first_name
        ws["O7"] = middle_name
        ws["Y7"] = last_name
    
    # Fill dual cells for deductions
    for json_key, cell_addresses in dual_cell_mapping.items():
        if json_key in form16_data:
            for cell_address in cell_addresses:
                ws[cell_address] = form16_data[json_key]
    
    # Try to remove existing file, if locked use different name
    try:
        if os.path.exists(output_file_path):
            os.remove(output_file_path)
        wb.save(output_file_path)
        print(f"Excel file filled and saved to: {output_file_path}")
    except PermissionError:
        # File is locked, use timestamp suffix
        import time
        timestamp = int(time.time())
        new_name = output_file_path.replace('.xlsx', f'_{timestamp}.xlsx')
        wb.save(new_name)
        print(f"Original file locked. Saved to: {new_name}")

if __name__ == "__main__":
    form16_json = "23d3cb64-a2b3-4d75-85c3-77a7923986a3_form16_parsed.json"
    aadhar_json = "aadhar_parsed.json"
    passbook_json = "passbook_parsed.json"
    excel_template = "itr_temp.xlsx"
    output_file = "filled_itr_complete.xlsx"
    
    fill_itr_excel(form16_json, aadhar_json, passbook_json, excel_template, output_file)

















