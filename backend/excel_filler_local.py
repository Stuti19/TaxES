import json
import os
from pathlib import Path
from openpyxl import load_workbook

class ExcelFiller:
    def __init__(self):
        self.base_dir = Path("taxes_files")
        self.parsed_dir = self.base_dir / "parsed"
        self.excel_dir = self.base_dir / "excel"
        
        # Create excel directory if it doesn't exist
        self.excel_dir.mkdir(parents=True, exist_ok=True)

    def fill_itr_excel(self, template_path=None):
        """Fill ITR Excel using parsed JSON files"""
        try:
            # Set template path
            if template_path is None:
                template_path = Path("itr_temp.xlsx")
                if not template_path.exists():
                    return {'status': 'error', 'message': 'Excel template not found'}
            
            # Load JSON data
            form16_path = self.parsed_dir / "form16_parsed.json"
            aadhar_path = self.parsed_dir / "aadhar_parsed.json"
            passbook_path = self.parsed_dir / "passbook_parsed.json"
            
            # Check if files exist
            if not form16_path.exists():
                return {'status': 'error', 'message': 'Form16 parsed data not found'}
            if not aadhar_path.exists():
                return {'status': 'error', 'message': 'Aadhar parsed data not found'}
            if not passbook_path.exists():
                return {'status': 'error', 'message': 'Passbook parsed data not found'}
            
            with open(form16_path, 'r') as f:
                form16_data = json.load(f)
            
            with open(aadhar_path, 'r') as f:
                aadhar_data = json.load(f)
            
            with open(passbook_path, 'r') as f:
                passbook_data = json.load(f)
            
            # Load Excel workbook
            print(f"Loading template from: {template_path}")
            wb = load_workbook(str(template_path))
            ws = wb.active
            print("Template loaded successfully")
            
            # Form 16 cell mapping
            form16_mapping = {
                "pan": "AN7",
                "employee_address": "O11",
                "gross_salary": "AO35",
                "salary_section_17_1": "AO36",
                "prerequisites_section_17_2": "AO37",
                "profits_section_17_3": "AO38",
                "total_exemption_section_10": "AO45",
                "standard_deduction_16_ia": "AO54",
                "entertainment_allowance_16_ii": "AO55",
                "tax_on_employment_16_iii": "AO56",
                "income_chargeable_salaries": "AO57",
                "gross_total_income": "AO93",
                "tax_on_total_income": "AO140",
                "rebate_87A": "AO141",
                "health_education_cess": "AO144",
                "relief_section_89": "AO146"
            }
            
            # Aadhar cell mapping
            aadhar_mapping = {
                "aadhar_number": "AN8",
                "dob": "AN11",
                "address": "O11"
            }
            
            # Deductions that need dual cells
            dual_cell_mapping = {
                "deduction_80C": ["AB96", "AN96"],
                "deduction_80CCC": ["AB97", "AN97"],
                "deduction_80CCD1": ["AB98", "AN98"],
                "deduction_80CCD1B": ["AB99", "AN99"],
                "deduction_80CCD2": ["AB101", "AN101"],
                "deduction_80D": ["AB102", "AN102"],
                "deduction_80E": ["AB111", "AN111"],
                "deduction_80G": ["AB115", "AN115"],
                "deduction_80TTA": ["AB122", "AN122"]
            }
            
            # Fill Form 16 data
            for json_key, cell_address in form16_mapping.items():
                if json_key in form16_data and form16_data[json_key]:
                    ws[cell_address] = form16_data[json_key]
            
            # Parse Aadhar data using Groq
            parsed_name = {}
            parsed_address = {}
            
            try:
                from groq_parser import GroqParser
                parser = GroqParser()
                
                if "name" in aadhar_data and aadhar_data["name"]:
                    parsed_name = parser.parse_name(aadhar_data["name"])
                    print(f"Parsed name: {parsed_name}")
                
                if "address" in aadhar_data and aadhar_data["address"]:
                    parsed_address = parser.parse_address(aadhar_data["address"])
                    print(f"Parsed address: {parsed_address}")
            except Exception as e:
                print(f"Groq parsing error: {e}")
            
            # Fill Aadhar data
            for json_key, cell_address in aadhar_mapping.items():
                if json_key in aadhar_data and aadhar_data[json_key]:
                    ws[cell_address] = aadhar_data[json_key]
            
            # Fill parsed name (prioritize Aadhar, fallback to passbook)
            if parsed_name:
                ws["E7"] = parsed_name.get("first_name", "")
                ws["O7"] = parsed_name.get("middle_name", "")
                ws["Y7"] = parsed_name.get("last_name", "")
            elif "name" in passbook_data and passbook_data["name"]:
                first_name, middle_name, last_name = self._parse_name(passbook_data["name"])
                ws["E7"] = first_name
                ws["O7"] = middle_name
                ws["Y7"] = last_name
            
            # Fill parsed address components
            if parsed_address:
                ws["E11"] = parsed_address.get("flat_door_block_no", "")  # Flat/Door/Block No
                ws["O11"] = parsed_address.get("premises_building_village", "")  # Building/Village name
                ws["E13"] = parsed_address.get("road_street_post_office", "")
                ws["W13"] = parsed_address.get("area_locality", "")
                ws["AN13"] = parsed_address.get("town_city_district", "")
                ws["E15"] = parsed_address.get("state", "")
                ws["AA15"] = parsed_address.get("pin_code", "")
            
            # Fill dual cells for deductions
            for json_key, cell_addresses in dual_cell_mapping.items():
                if json_key in form16_data and form16_data[json_key]:
                    for cell_address in cell_addresses:
                        ws[cell_address] = form16_data[json_key]
            
            # Calculate and fill derived tax fields
            self._fill_calculated_tax_fields(ws, form16_data)
            
            # Save to excel folder
            output_path = self.excel_dir / "filled_itr.xlsx"
            print(f"Saving Excel to: {output_path}")
            
            # Remove existing file if it exists
            if output_path.exists():
                output_path.unlink()
            
            wb.save(str(output_path))
            print("Excel saved successfully")
            
            return {
                'status': 'success',
                'message': 'Excel file generated successfully',
                'file_path': str(output_path)
            }
            
        except Exception as e:
            print(f"Excel generation error: {str(e)}")
            import traceback
            traceback.print_exc()
            return {
                'status': 'error',
                'message': f'Error generating Excel: {str(e)}'
            }

    def _fill_calculated_tax_fields(self, ws, form16_data):
        """Calculate and fill derived tax fields"""
        try:
            # Get values with default 0 if not present or not numeric
            tax_on_total_income = self._get_numeric_value(form16_data, "tax_on_total_income")
            rebate_87A = self._get_numeric_value(form16_data, "rebate_87A")
            health_education_cess = self._get_numeric_value(form16_data, "health_education_cess")
            relief_section_89 = self._get_numeric_value(form16_data, "relief_section_89")
            
            # Calculate tax_payable_after_rebate = tax_on_total_income - rebate_87A
            tax_payable_after_rebate = tax_on_total_income - rebate_87A
            ws["AO142"] = tax_payable_after_rebate
            
            # Calculate total_tax_and_cess = tax_payable_after_rebate + health_education_cess
            total_tax_and_cess = tax_payable_after_rebate + health_education_cess
            ws["AO145"] = total_tax_and_cess
            
            # Calculate balance_tax_after_relief = total_tax_and_cess - relief_section_89
            balance_tax_after_relief = total_tax_and_cess - relief_section_89
            ws["AO148"] = balance_tax_after_relief
            
            print(f"Calculated tax fields:")
            print(f"  Tax payable after rebate (AO142): {tax_payable_after_rebate}")
            print(f"  Total tax and cess (AO145): {total_tax_and_cess}")
            print(f"  Balance tax after relief (AO148): {balance_tax_after_relief}")
            
        except Exception as e:
            print(f"Error calculating tax fields: {e}")
    
    def _get_numeric_value(self, data, key):
        """Get numeric value from data, return 0 if not present or not numeric"""
        try:
            value = data.get(key, 0)
            if isinstance(value, (int, float)):
                return float(value)
            elif isinstance(value, str):
                # Try to convert string to float
                return float(value) if value.strip() else 0.0
            else:
                return 0.0
        except (ValueError, TypeError):
            return 0.0
    
    def _parse_name(self, full_name):
        """Parse name into first, middle, last"""
        if not full_name or full_name.lower() == "no":
            return "", "", ""
        
        # Remove common prefixes and clean
        name_parts = full_name.replace("MR ", "").replace("MRS ", "").replace("MS ", "").strip().split()
        
        if len(name_parts) == 1:
            return name_parts[0], "", ""
        elif len(name_parts) == 2:
            return name_parts[0], "", name_parts[1]
        else:
            return name_parts[0], " ".join(name_parts[1:-1]), name_parts[-1]